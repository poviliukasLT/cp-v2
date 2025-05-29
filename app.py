import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.formula.translate import Translator
from io import BytesIO
from PIL import Image
from datetime import datetime
import pytz
import unicodedata

st.set_page_config(page_title="Pasiūlymų generatorius", layout="wide")

st.markdown("""
    <style>
        body, .stApp {
            background-color: #ffffff;
        }
        .centered-logo {
            display: flex;
            justify-content: center;
            margin-bottom: 10px;
        }
    </style>
""", unsafe_allow_html=True)

logo = Image.open("logo-red.png")
st.markdown("<div style='text-align: center;'>", unsafe_allow_html=True)
st.image(logo, width=300)
st.markdown("</div>", unsafe_allow_html=True)

st.title("📦 Pasiūlymų kūrimo įrankis v4.6-patched")

if 'pasirinktos_eilutes' not in st.session_state:
    st.session_state.pasirinktos_eilutes = []
if 'pasirinktu_failu_pavadinimai' not in st.session_state:
    st.session_state.pasirinktu_failu_pavadinimai = []
if 'pasirinktu_formuliu_info' not in st.session_state:
    st.session_state.pasirinktu_formuliu_info = []

rename_rules = {
    "Sweets": ["", "Product code", "Product name", "Purchasing price", "Label",
               "Price with costs", "Target Margin", "Target offer", "VAT",
               "Offer with VAT", "RSP MIN", "RSP MAX", "Margin RSP MIN", "Margin RSP MAX",
               "", "Target Margin", "Target offer"],
    "Snacks_": ["", "Product code", "Product name", "Purchasing price", "Label",
                "Price with costs", "Target Margin", "Target offer", "VAT",
                "Offer with VAT", "RSP MIN", "RSP MAX", "Margin RSP MIN", "Margin RSP MAX",
                "", "Target Margin", "Target offer"],
    "Groceries": ["", "Product code", "Product name", "Purchasing price", "Label",
                  "Price with costs", "Target Margin", "Target offer", "VAT",
                  "Offer with VAT", "RSP MIN", "RSP MAX", "Margin RSP MIN", "Margin RSP MAX",
                  "", "Target Margin", "Target offer"],
    "beverages": ["Country", "Product code", "Product name", "Purchasing price", "Label",
                  "Deposit (if needed)", "Sugar Tax", "Price with costs", "Target Margin",
                  "Target offer", "VAT", "Offer with VAT", "RSP MIN", "RSP MAX",
                  "Margin RSP MIN", "Margin RSP MAX", "Target Margin", "Target offer",
                  "", "AS OF 2025", "CAN up to 0,33l", "CAN over 0,33",
                  "PET up to 0,75l", "PET over 0,75l", "GLASS up to 0,5l", "GLASS over 0,5l"]
}

proc_format_map = {
    "Sweets": ["Target Margin", "VAT", "Margin RSP MIN", "Margin RSP MAX"],
    "Snacks_": ["Target Margin", "VAT", "Margin RSP MIN", "Margin RSP MAX"],
    "Groceries": ["Target Margin", "VAT", "Margin RSP MIN", "Margin RSP MAX"],
    "beverages": ["Target Margin", "VAT", "Margin RSP MIN", "Margin RSP MAX", "Target margin"]
}

def normalize(text):
    if not isinstance(text, str):
        return ""
    text = unicodedata.normalize("NFKD", text)
    return "".join(text.lower().strip().replace("\u00a0", "").split())

@st.cache_data
def extract_rows_with_metadata(file):
    wb = load_workbook(file, data_only=False)
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        formulas = []
        for row in ws.iter_rows(values_only=False):
            row_data = []
            formula_row = []
            for cell in row:
                if cell.data_type == 'f':
                    formula_row.append((cell.coordinate, f"={cell.value}"))
                    row_data.append(f"={cell.value}")
                else:
                    formula_row.append(None)
                    row_data.append(cell.value)
            rows.append(row_data)
            formulas.append(formula_row)
        data[f"{file.name} -> {sheet_name}"] = (rows, formulas, file.name.split(".")[0])
    return data

uploaded_files = st.file_uploader("📁 Įkelkite Excel failus:", type="xlsx", accept_multiple_files=True)

if uploaded_files:
    all_data = {}
    for file in uploaded_files:
        extracted = extract_rows_with_metadata(file)
        all_data.update(extracted)

    pasirinkimas = st.selectbox("Pasirinkite failą ir lapą:", list(all_data.keys()))
    rows, formulas, failo_pav = all_data[pasirinkimas]
    df = pd.DataFrame(rows)
    st.dataframe(df.head(100))

    pasirinktos = st.multiselect("✅ Pasirinkite eilučių numerius:", df.index)
    if st.button("➕ Pridėti pažymėtas"):
        for i in pasirinktos:
            eilute = df.iloc[i].tolist()
            if eilute not in st.session_state.pasirinktos_eilutes:
                st.session_state.pasirinktos_eilutes.append(eilute)
                st.session_state.pasirinktu_failu_pavadinimai.append(failo_pav)
                st.session_state.pasirinktu_formuliu_info.append(formulas[i])
            else:
                st.warning(f"Eilutė #{i} jau pridėta ir nebus dubliuojama.")

st.subheader("🧠 Atmintis")
if not st.session_state.pasirinktos_eilutes:
    st.info("Nėra pasirinkimų.")
else:
    df_memory = pd.DataFrame(st.session_state.pasirinktos_eilutes)
    st.dataframe(df_memory)
    pasirinkti_salinimui = st.multiselect("🗑️ Pažymėkite eilutes pašalinimui:", df_memory.index)
    col1, col2 = st.columns(2)
    if col1.button("❌ Pašalinti pažymėtas"):
        st.session_state.pasirinktos_eilutes = [r for i, r in enumerate(st.session_state.pasirinktos_eilutes) if i not in pasirinkti_salinimui]
        st.session_state.pasirinktu_failu_pavadinimai = [n for i, n in enumerate(st.session_state.pasirinktu_failu_pavadinimai) if i not in pasirinkti_salinimui]
        st.session_state.pasirinktu_formuliu_info = [f for i, f in enumerate(st.session_state.pasirinktu_formuliu_info) if i not in pasirinkti_salinimui]
    if col2.button("🧹 Išvalyti viską"):
        st.session_state.pasirinktos_eilutes = []
        st.session_state.pasirinktu_failu_pavadinimai = []
        st.session_state.pasirinktu_formuliu_info = []
        st.rerun()

if st.session_state.pasirinktos_eilutes and st.session_state.pasirinktu_failu_pavadinimai:
    if st.button("📅 Eksportuoti su koreguotomis formulėmis"):
        wb = Workbook()
        ws = wb.active

        grouped = {}
        for i, failas in enumerate(st.session_state.pasirinktu_failu_pavadinimai):
            if failas not in grouped:
                grouped[failas] = []
            grouped[failas].append((st.session_state.pasirinktos_eilutes[i], st.session_state.pasirinktu_formuliu_info[i]))

        row_pointer = 1
        is_beverages_included = any(f.lower().startswith("beverages") for f in grouped)

        for failas, eilutes_info in grouped.items():
            matching_key = None
            for key in rename_rules:
                if failas.lower().startswith(key.lower()):
                    matching_key = key
                    break

            header = rename_rules.get(matching_key, [f"Column {i+1}" for i in range(len(eilutes_info[0][0]))])
            proc_columns = proc_format_map.get(matching_key, [])
            proc_format_names = [normalize(n) for n in proc_columns]
            proc_format_indexes = [idx for idx, name in enumerate(header) if normalize(name) in proc_format_names]

            header_row = header[:]
            if matching_key != "beverages" and is_beverages_included:
                header_row = header_row[:5] + ["", ""] + header_row[5:]

            for col_idx, val in enumerate(header_row):
                ws.cell(row=row_pointer, column=col_idx + 1).value = val
            row_pointer += 1

            for row_data, formula_row in eilutes_info:
                adjusted_row = row_data[:]
                adjusted_formula = formula_row[:]
                if matching_key != "beverages" and is_beverages_included:
                    adjusted_row = adjusted_row[:5] + [None, None] + adjusted_row[5:]
                    adjusted_formula = adjusted_formula[:5] + [None, None] + adjusted_formula[5:]

                for col_idx, value in enumerate(adjusted_row):
                    export_cell = ws.cell(row=row_pointer, column=col_idx + 1)
                    formula_info = adjusted_formula[col_idx] if col_idx < len(adjusted_formula) else None
                    if formula_info:
                        original_coord, formula_text = formula_info
                        translated = Translator(formula_text, origin=original_coord).translate_formula(export_cell.coordinate)
                        export_cell.value = translated
                    else:
                        export_cell.value = value
                    if col_idx in proc_format_indexes:
                        export_cell.number_format = "0.00%"
                row_pointer += 1

            row_pointer += 1

        lt_tz = pytz.timezone("Europe/Vilnius")
        now_str = datetime.now(lt_tz).strftime("%Y-%m-%d_%H-%M")
        output = BytesIO()
        wb.save(output)
        st.download_button(
            label="📅 Atsisiųsti su koreguotomis formulėmis",
            data=output.getvalue(),
            file_name=f"pasiulymas_{now_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
